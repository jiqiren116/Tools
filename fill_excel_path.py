import re
import openpyxl

# ====================== 手动配置区（请根据实际情况修改）======================
EXCEL_PATH = r"D:\Cmy\Code\Tools\优博讯SQ51QV代码推送记录(1) - 副本.xlsx"  #  Excel文件路径
TXT_PATH = r"D:\Cmy\Code\Tools\changegit_20260210.txt"              # TXT文件路径
SHEET_NAME = "20260210"                               # 要处理的Sheet名
SKIP_HEADER = False                                   # 关键：是否跳过第一行表头【True=跳过，False=不跳过（无表头时设为False）】
# ==============================================================================

def parse_txt_to_dict(txt_path):
    """解析TXT，生成{name: path}映射字典，name做去空白处理"""
    name_path_dict = {}
    # 正则匹配project的name和path，兼容单/双引号、大小写
    pattern = re.compile(r'<project\s+name=["\'](.*?)["\']\s+path=["\'](.*?)["\']', re.IGNORECASE)
    
    try:
        with open(txt_path, "r", encoding="utf-8") as f:
            matches = pattern.findall(f.read())
            for name, path in matches:
                clean_name = name.strip()  # 去除name首尾空格/换行
                name_path_dict[clean_name] = path.strip()
        print(f"✅ TXT解析完成，共提取 {len(name_path_dict)} 个映射关系")
        return name_path_dict
    except FileNotFoundError:
        print(f"❌ 错误：未找到TXT文件 {txt_path}")
        exit(1)
    except Exception as e:
        print(f"❌ TXT解析失败：{str(e)}")
        exit(1)

def fill_excel_d_column(excel_path, sheet_name, name_path_dict, skip_header):
    """填充Excel D列，支持控制是否跳过表头，C列值自动去空白"""
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"❌ 错误：未找到Excel文件 {excel_path}")
        exit(1)
    except Exception as e:
        print(f"❌ Excel加载失败：{str(e)}")
        exit(1)

    if sheet_name not in wb.sheetnames:
        print(f"❌ 错误：无Sheet {sheet_name}，现有Sheet：{wb.sheetnames}")
        wb.close()
        exit(1)

    ws = wb[sheet_name]
    fill_count = 0
    max_row = ws.max_row
    # 确定开始行：跳过表头则从2行开始，不跳过则从1行开始
    start_row = 2 if skip_header else 1

    # 遍历所有数据行，C列（3列）值自动去空白，避免空格导致匹配失败
    for row in range(start_row, max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        c_value = c_cell.value
        if not c_value:
            continue
        # 关键：对C列值做去空白处理，匹配更精准
        clean_c_value = str(c_value).strip()
        if clean_c_value in name_path_dict:
            ws.cell(row=row, column=4, value=name_path_dict[clean_c_value])
            fill_count += 1
            print(f"✅ 第{row}行匹配成功：{clean_c_value} -> {name_path_dict[clean_c_value]}")
        else:
            print(f"⚠️  第{row}行无匹配结果：{clean_c_value}")

    # 保存文件并提示
    try:
        wb.save(excel_path)
        print(f"\n===== 处理完成 =====\n✅ 共成功填充 {fill_count} 行数据\n✅ 文件已保存：{excel_path}")
    except PermissionError:
        print(f"❌ 错误：Excel文件被占用，请关闭后重新运行")
    except Exception as e:
        print(f"❌ Excel保存失败：{str(e)}")
    finally:
        wb.close()

if __name__ == "__main__":
    print("===== 开始处理Excel填充 =====")
    name_path_map = parse_txt_to_dict(TXT_PATH)
    fill_excel_d_column(EXCEL_PATH, SHEET_NAME, name_path_map, SKIP_HEADER)